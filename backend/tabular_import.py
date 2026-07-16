"""Import de fichiers tabulaires (xlsx / csv / tsv) avec détection de colonnes.

Objectif : accepter directement un vrai listing (ex. export Excel à 8 colonnes)
sans que l'utilisateur ait à le retravailler. On :
  1. lit le fichier et détecte la ligne d'en-tête,
  2. liste les colonnes avec des échantillons,
  3. devine la colonne « numéro » (score = % de valeurs qui sont des numéros FR
     valides) et propose une colonne « libellé »,
  4. laisse l'appelant choisir le mapping, puis produit les numéros normalisés en
     conservant TOUTES les colonnes du fichier dans `extras` (affichage cockpit).

Formats : .xlsx (openpyxl), .csv/.tsv/.txt (module csv, séparateur détecté).
"""
from __future__ import annotations

import csv
import io
from typing import Any

from phone_numbers import normalize_number


MAX_PREVIEW_ROWS = 5


# --- Lecture brute : fichier → grille de cellules (liste de lignes) -----------

def read_xlsx(data: bytes) -> list[list[str]]:
    """Lit la première feuille non vide d'un .xlsx en grille de chaînes."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    # Choisir la feuille avec le plus de lignes non vides.
    best_ws = None
    best_count = -1
    for ws in wb.worksheets:
        count = sum(1 for row in ws.iter_rows(values_only=True) if any(c is not None for c in row))
        if count > best_count:
            best_count = count
            best_ws = ws
    grid: list[list[str]] = []
    if best_ws is not None:
        for row in best_ws.iter_rows(values_only=True):
            grid.append(["" if c is None else str(c).strip() for c in row])
    wb.close()
    return grid


def read_delimited(data: bytes) -> list[list[str]]:
    """Lit un CSV/TSV en grille. Détecte le séparateur (; , tab)."""
    text = _decode(data)
    # Détecter le séparateur sur les premières lignes non vides.
    sample = "\n".join([ln for ln in text.splitlines() if ln.strip()][:10])
    delimiter = ";"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        delimiter = dialect.delimiter
    except csv.Error:
        # Repli : compter les occurrences.
        counts = {d: sample.count(d) for d in [";", ",", "\t"]}
        delimiter = max(counts, key=counts.get) if any(counts.values()) else ";"
    grid: list[list[str]] = []
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    for row in reader:
        grid.append([c.strip() for c in row])
    return grid


def _decode(data: bytes) -> str:
    """Décode des octets texte en tolérant l'encodage (BOM, UTF-8, latin-1)."""
    if data[:3] == b"\xef\xbb\xbf":
        return data.decode("utf-8-sig")
    if data[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return data.decode("utf-16")
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("latin-1", errors="replace")


def read_file(filename: str, data: bytes) -> list[list[str]]:
    """Aiguille selon l'extension."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return read_xlsx(data)
    return read_delimited(data)


# --- Détection de l'en-tête et des colonnes -----------------------------------

def _looks_like_fr_number(value: str) -> bool:
    if not value:
        return False
    try:
        normalize_number(value)
        return True
    except ValueError:
        return False


def _find_header_row(grid: list[list[str]]) -> int:
    """Trouve l'indice de la ligne d'en-tête.

    Heuristique : la première ligne qui a >= 2 cellules texte non vides ET dont
    AUCUNE cellule n'est un numéro FR (les vraies données commencent après).
    On saute les lignes de titre (1 seule cellule remplie).

    Renvoie -1 si AUCUN en-tête n'est trouvé (fichier sans ligne d'en-tête : la
    première ligne de données contient déjà des numéros). Dans ce cas l'appelant
    génère des noms de colonnes génériques et traite toutes les lignes en données.
    """
    for i, row in enumerate(grid[:15]):
        non_empty = [c for c in row if c.strip()]
        if len(non_empty) < 2:
            continue
        has_number = any(_looks_like_fr_number(c) for c in row)
        if not has_number:
            return i
    # Aucune ligne d'en-tête crédible : pas d'en-tête (données dès la 1re ligne).
    return -1


def analyze(grid: list[list[str]]) -> dict[str, Any]:
    """Analyse la grille : en-tête, colonnes, échantillons, colonne numéro devinée.

    Renvoie :
    {
      "header_row": int,
      "columns": [
        {"index": 0, "name": "N°", "samples": [...],
         "number_score": 0.0, "is_number_guess": False, "is_label_guess": False},
        ...
      ],
      "data_row_count": int,
      "number_col_guess": <index|None>,
      "label_col_guess": <index|None>,
      "single_column": bool   # fichier à 1 colonne = collage simple
    }
    """
    if not grid:
        return {"header_row": 0, "columns": [], "data_row_count": 0,
                "number_col_guess": None, "label_col_guess": None, "single_column": True}

    header_idx = _find_header_row(grid)
    has_header = header_idx >= 0
    header = grid[header_idx] if has_header else []
    # Sans en-tête : toutes les lignes non vides sont des données.
    first_data = (header_idx + 1) if has_header else 0
    data_rows = [r for r in grid[first_data:] if any(c.strip() for c in r)]
    ncols = max((len(r) for r in grid), default=0)

    columns: list[dict[str, Any]] = []
    for col in range(ncols):
        name = header[col].strip() if (has_header and col < len(header)) else ""
        if not name:
            name = f"Colonne {col + 1}"
        values = [r[col].strip() for r in data_rows if col < len(r) and r[col].strip()]
        samples = values[:MAX_PREVIEW_ROWS]
        # Score numéro : proportion de valeurs qui sont des numéros FR valides.
        valid = sum(1 for v in values if _looks_like_fr_number(v))
        score = (valid / len(values)) if values else 0.0
        columns.append({
            "index": col,
            "name": name,
            "samples": samples,
            "number_score": round(score, 3),
            "is_number_guess": False,
            "is_label_guess": False,
        })

    # Colonne numéro = meilleur score (si >= 50%).
    number_col = None
    best = 0.5
    for c in columns:
        if c["number_score"] >= best:
            best = c["number_score"]
            number_col = c["index"]
    # Si aucune colonne ne dépasse 50%, prendre la meilleure non nulle.
    if number_col is None:
        nonzero = [c for c in columns if c["number_score"] > 0]
        if nonzero:
            number_col = max(nonzero, key=lambda c: c["number_score"])["index"]

    # Colonne libellé = une colonne texte (score numéro faible) avec un nom
    # évocateur (site, libellé, nom, destination) sinon la 1re colonne texte.
    label_col = None
    label_keywords = ("site", "libell", "label", "nom", "destination", "client", "lieu")
    text_cols = [c for c in columns if c["index"] != number_col and c["number_score"] < 0.3]
    for c in text_cols:
        if any(k in c["name"].lower() for k in label_keywords):
            label_col = c["index"]
            break
    if label_col is None and text_cols:
        # Première colonne texte non triviale (pas un simple compteur).
        for c in text_cols:
            vals = c["samples"]
            # Éviter les colonnes de numérotation (1,2,3…).
            if not all(v.isdigit() and len(v) < 4 for v in vals if v):
                label_col = c["index"]
                break

    for c in columns:
        c["is_number_guess"] = (c["index"] == number_col)
        c["is_label_guess"] = (c["index"] == label_col)

    return {
        "header_row": header_idx,
        "columns": columns,
        "data_row_count": len(data_rows),
        "number_col_guess": number_col,
        "label_col_guess": label_col,
        "single_column": ncols <= 1,
    }


# --- Construction des numéros avec mapping -------------------------------------

def build_numbers(
    grid: list[list[str]],
    header_row: int,
    number_col: int,
    label_col: int | None,
    existing: set[str] | None = None,
) -> dict[str, Any]:
    """Applique le mapping choisi et produit valides / rejets / doublons.

    Conserve TOUTES les colonnes de chaque ligne dans `extras` (nom_colonne → valeur).
    Structure de retour identique à ImportResult.as_dict() + extras dans chaque valide.
    """
    existing = existing or set()
    # header_row == -1 : pas d'en-tête → noms génériques, données dès la 1re ligne.
    has_header = header_row >= 0
    header = grid[header_row] if (has_header and header_row < len(grid)) else []
    col_names = [
        (header[i].strip() if (has_header and i < len(header) and header[i].strip())
         else f"Colonne {i + 1}")
        for i in range(max((len(r) for r in grid), default=0))
    ]

    valid: list[dict[str, Any]] = []
    rejected: list[dict[str, str]] = []
    duplicates: list[dict[str, str]] = []
    seen: set[str] = set()

    for row in grid[(header_row + 1) if has_header else 0:]:
        if not any(c.strip() for c in row):
            continue
        number_raw = row[number_col].strip() if number_col < len(row) else ""
        if not number_raw:
            # Ligne sans numéro dans la colonne choisie : rejet informatif.
            first = next((c for c in row if c.strip()), "")
            if first:
                rejected.append({"raw": first, "reason": "Pas de numéro dans la colonne choisie"})
            continue
        try:
            norm = normalize_number(number_raw)
        except ValueError as e:
            rejected.append({"raw": number_raw, "reason": str(e)})
            continue

        e164 = norm["e164"]
        if e164 in existing or e164 in seen:
            duplicates.append({"raw": number_raw, "e164": e164})
            continue
        seen.add(e164)

        label = ""
        if label_col is not None and label_col < len(row):
            label = row[label_col].strip()

        # extras = toutes les colonnes non vides (hors la colonne numéro brute).
        extras: dict[str, str] = {}
        for i, cell in enumerate(row):
            val = cell.strip() if i < len(row) else ""
            if not val:
                continue
            key = col_names[i] if i < len(col_names) else f"Colonne {i + 1}"
            extras[key] = val

        valid.append({
            "raw": number_raw,
            "e164": e164,
            "national": norm["national"],
            "label": label,
            "extras": extras,
        })

    return {
        "valid": valid,
        "rejected": rejected,
        "duplicates": duplicates,
        "counts": {"valid": len(valid), "rejected": len(rejected), "duplicates": len(duplicates)},
    }
