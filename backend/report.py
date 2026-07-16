"""Rapport comparatif : vue croisée par numéro + classification automatique + exports.

Classification (CLAUDE.md §6.5), par numéro, sur l'ensemble des passes de la campagne :
  - OK + OK                          → Conforme
  - OK d'un côté, NOK de l'autre     → Routage inter-opérateurs suspect
  - NOK + NOK                        → Portage KO
  - Verdict sur une seule passe      → Partiel (mono-SIM ou 2e passe non faite)
  - SKIP                             → Non testé (sur cette passe)

Le rapport reste cohérent et exploitable avec une seule passe.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any

import database as db


# Catégories (clés stables pour l'UI + libellés FR).
CAT_CONFORME = "conforme"
CAT_ROUTAGE = "routage_suspect"
CAT_KO = "portage_ko"
CAT_PARTIEL = "partiel"
CAT_NON_TESTE = "non_teste"

CAT_LABELS = {
    CAT_CONFORME: "Conforme",
    CAT_ROUTAGE: "⚠ Routage inter-opérateurs suspect",
    CAT_KO: "✖ Portage KO",
    CAT_PARTIEL: "Partiel",
    CAT_NON_TESTE: "Non testé",
}


def _classify(verdicts: list[str | None]) -> str:
    """Classe un numéro à partir de la liste de ses verdicts (un par passe).

    verdicts contient les verdicts effectifs des passes existantes (OK/NOK/SKIP/None).
    None = passe existante mais numéro pas encore traité → traité comme absence.
    """
    # On ne considère que les verdicts « décisifs » OK/NOK pour la conformité.
    decisive = [v for v in verdicts if v in ("OK", "NOK")]
    has_skip = any(v == "SKIP" for v in verdicts)

    if len(decisive) == 0:
        # Aucun verdict décisif : soit uniquement des SKIP, soit rien.
        return CAT_NON_TESTE if has_skip else CAT_NON_TESTE

    oks = decisive.count("OK")
    noks = decisive.count("NOK")

    if len(decisive) == 1:
        # Une seule passe décisive → Partiel (qu'elle soit OK ou NOK).
        return CAT_PARTIEL

    # Au moins deux passes décisives.
    if oks >= 1 and noks >= 1:
        return CAT_ROUTAGE
    if noks == len(decisive):
        return CAT_KO
    # Tous OK.
    return CAT_CONFORME


def build_report(campaign_id: int) -> dict[str, Any]:
    """Construit le rapport croisé d'une campagne.

    Structure :
    {
      "campaign": {...},
      "runs": [{id, sim_operator, sim_slot, started_at, status}, ...],
      "rows": [
        {number_id, national, e164, label,
         "cells": {run_id: {verdict, duration_s, comment}},
         "category": "...", "category_label": "..."}
      ],
      "summary": {category: count, ...},
      "total": N
    }
    """
    campaign = db.get_campaign(campaign_id)
    runs = db.list_runs(campaign_id)
    # Ordre chronologique des passes pour l'affichage (plus ancienne à gauche).
    runs_sorted = sorted(runs, key=lambda r: r["id"])
    numbers = db.list_numbers(campaign_id)

    rows: list[dict[str, Any]] = []
    summary: dict[str, int] = {c: 0 for c in CAT_LABELS}

    for num in numbers:
        cells: dict[int, dict[str, Any]] = {}
        verdicts: list[str | None] = []
        for run in runs_sorted:
            call = db.get_call_for(run["id"], num["id"])
            if call:
                cells[run["id"]] = {
                    "verdict": call["verdict"],
                    "duration_s": call["duration_s"],
                    "comment": call["comment"] or "",
                }
                verdicts.append(call["verdict"])
            else:
                cells[run["id"]] = {"verdict": None, "duration_s": None, "comment": ""}
                verdicts.append(None)

        category = _classify(verdicts)
        summary[category] += 1
        rows.append({
            "number_id": num["id"],
            "national": num["national"],
            "e164": num["e164"],
            "label": num["label"] or "",
            "cells": cells,
            "category": category,
            "category_label": CAT_LABELS[category],
        })

    return {
        "campaign": campaign,
        "runs": [
            {
                "id": r["id"],
                "sim_operator": r["sim_operator"],
                "sim_slot": r["sim_slot"],
                "sim_subid": r["sim_subid"],
                "started_at": r["started_at"],
                "status": r["status"],
            }
            for r in runs_sorted
        ],
        "rows": rows,
        "summary": summary,
        "summary_labels": CAT_LABELS,
        "total": len(numbers),
    }


# --- Exports ------------------------------------------------------------------

def _run_col_prefix(run: dict[str, Any]) -> str:
    op = run.get("sim_operator") or "SIM"
    date = (run.get("started_at") or "")[:10]
    return f"{op} {date}".strip()


def _safe_filename(name: str) -> str:
    keep = "".join(c if c.isalnum() or c in " -_" else "_" for c in (name or "campagne"))
    return keep.strip().replace(" ", "_") or "campagne"


def export_filename(campaign_name: str, ext: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    return f"rapport_{_safe_filename(campaign_name)}_{stamp}.{ext}"


def export_csv(campaign_id: int) -> tuple[str, str]:
    """Renvoie (nom_fichier, contenu_csv). UTF-8, séparateur ';'. Une ligne par numéro."""
    report = build_report(campaign_id)
    runs = report["runs"]
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", lineterminator="\r\n")

    header = ["Numéro (national)", "E.164", "Libellé"]
    for run in runs:
        prefix = _run_col_prefix(run)
        header += [f"{prefix} - verdict", f"{prefix} - durée (s)", f"{prefix} - commentaire"]
    header += ["Classification"]
    writer.writerow(header)

    for row in report["rows"]:
        line = [row["national"], row["e164"], row["label"]]
        for run in runs:
            cell = row["cells"].get(run["id"], {})
            line += [
                cell.get("verdict") or "",
                cell.get("duration_s") if cell.get("duration_s") is not None else "",
                cell.get("comment") or "",
            ]
        line += [row["category_label"]]
        writer.writerow(line)

    name = export_filename(report["campaign"]["name"] if report["campaign"] else "campagne", "csv")
    # BOM UTF-8 pour qu'Excel ouvre correctement les accents.
    return name, "﻿" + output.getvalue()


def export_xlsx(campaign_id: int) -> tuple[str, bytes]:
    """Renvoie (nom_fichier, contenu_xlsx_bytes) via openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    report = build_report(campaign_id)
    runs = report["runs"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"

    header = ["Numéro (national)", "E.164", "Libellé"]
    for run in runs:
        prefix = _run_col_prefix(run)
        header += [f"{prefix} - verdict", f"{prefix} - durée (s)", f"{prefix} - commentaire"]
    header += ["Classification"]
    ws.append(header)

    # Style en-tête.
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Couleurs par catégorie.
    cat_fill = {
        CAT_CONFORME: "D1FAE5",   # vert clair
        CAT_ROUTAGE: "FEF3C7",    # ambre
        CAT_KO: "FEE2E2",         # rouge clair
        CAT_PARTIEL: "E5E7EB",    # gris
        CAT_NON_TESTE: "F3F4F6",  # gris très clair
    }

    for row in report["rows"]:
        line = [row["national"], row["e164"], row["label"]]
        for run in runs:
            cell = row["cells"].get(run["id"], {})
            line += [
                cell.get("verdict") or "",
                cell.get("duration_s") if cell.get("duration_s") is not None else "",
                cell.get("comment") or "",
            ]
        line += [row["category_label"]]
        ws.append(line)
        # Colorer la cellule de classification.
        last_cell = ws.cell(row=ws.max_row, column=len(line))
        fill_color = cat_fill.get(row["category"], "FFFFFF")
        last_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Largeurs de colonnes.
    widths = [18, 16, 20] + [14, 12, 24] * len(runs) + [32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    # Feuille de synthèse.
    ws2 = wb.create_sheet("Synthèse")
    ws2.append(["Catégorie", "Nombre"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for cat, label in report["summary_labels"].items():
        ws2.append([label, report["summary"].get(cat, 0)])
    ws2.append(["Total", report["total"]])
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)
    name = export_filename(report["campaign"]["name"] if report["campaign"] else "campagne", "xlsx")
    return name, buffer.getvalue()
